"""Checked exports for official result workbooks and Origin-ready tables."""

from __future__ import annotations

from pathlib import Path
import shutil
from collections.abc import Mapping
from numbers import Real

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font    


def _as_frame(data, column_order=None):
    frame = data.copy() if isinstance(data, pd.DataFrame) else pd.DataFrame(data)
    if column_order is not None:
        missing = [name for name in column_order if name not in frame.columns]
        if missing:
            raise ValueError(f"missing columns: {missing}")
        frame = frame.loc[:, list(column_order)]
    return frame


def _inf_count(frame):
    """Count positive/negative infinity without coercing text columns."""
    numeric = frame.select_dtypes(include=[np.number])
    return int(np.isinf(numeric.to_numpy(dtype=float)).sum()) if not numeric.empty else 0


def write_excel_checked(path, sheets, *, column_orders=None, decimals=None, allow_nan=False):
    """Write one or more result sheets, style them, reread, and verify shape/order/NaN."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if isinstance(sheets, pd.DataFrame):
        sheets = {"Results": sheets}
    elif isinstance(sheets, Mapping):
        is_multisheet = bool(sheets) and all(isinstance(value, (pd.DataFrame, Mapping)) for value in sheets.values())
        if not is_multisheet:
            sheets = {"Results": sheets}
    else:
        sheets = {"Results": sheets}
    frames = {}
    for name, data in sheets.items():
        order = None if column_orders is None else column_orders.get(name)
        frame = _as_frame(data, order)
        if not allow_nan and frame.isna().any().any():
            raise ValueError(f"sheet {name!r} contains NaN")
        if _inf_count(frame):
            raise ValueError(f"sheet {name!r} contains Inf")
        frames[str(name)[:31]] = frame
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in frames.items():
            frame.to_excel(writer, sheet_name=name, index=False, float_format=None if decimals is None else f"%.{decimals}f")
    workbook = load_workbook(path)
    for worksheet in workbook.worksheets:
        for cell in worksheet[1]:
            cell.font = Font(name="Arial", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for column in worksheet.columns:
            letter = column[0].column_letter
            worksheet.column_dimensions[letter].width = min(30, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            for cell in column[1:]:
                cell.font = Font(name="Arial")
                if decimals is not None and isinstance(cell.value, float):
                    cell.number_format = "0." + "0" * int(decimals)
    workbook.save(path)
    verification = verify_excel(path, expected={name: frame for name, frame in frames.items()}, allow_nan=allow_nan)
    if not verification["valid"]:
        raise RuntimeError(f"Excel verification failed: {verification['errors']}")
    return verification


def verify_excel(path, *, expected=None, allow_nan=False):
    path = Path(path)
    errors = []
    if not path.exists() or path.stat().st_size == 0:
        return {"valid": False, "errors": ["file missing or empty"], "sheets": {}}
    actual = pd.read_excel(path, sheet_name=None)
    details = {
        name: {
            "rows": len(frame),
            "columns": len(frame.columns),
            "column_order": list(frame.columns),
            "nan_count": int(frame.isna().sum().sum()),
            "inf_count": _inf_count(frame),
        }
        for name, frame in actual.items()
    }
    if not allow_nan:
        errors.extend(f"{name}: contains NaN" for name, frame in actual.items() if frame.isna().any().any())
    errors.extend(f"{name}: contains Inf" for name, frame in actual.items() if _inf_count(frame))
    if expected is not None:
        for name, frame in expected.items():
            if name not in actual:
                errors.append(f"missing sheet {name}")
            elif actual[name].shape != frame.shape:
                errors.append(f"{name}: shape {actual[name].shape} != {frame.shape}")
            elif list(actual[name].columns) != list(frame.columns):
                errors.append(f"{name}: column order changed")
    return {"valid": not errors, "errors": errors, "sheets": details, "path": str(path)}


def export_origin_table(path, data, *, x_column=None, column_order=None, metadata=None):
    """Export one small, plot-specific Origin workbook with X first and optional Notes sheet."""
    frame = _as_frame(data, column_order)
    if x_column is not None:
        if x_column not in frame.columns:
            raise ValueError(f"x column {x_column!r} is missing")
        frame = frame[[x_column] + [column for column in frame.columns if column != x_column]]
    sheets = {"Data": frame}
    if metadata:
        sheets["Notes"] = pd.DataFrame({"item": list(metadata), "description": list(metadata.values())})
    return write_excel_checked(path, sheets, allow_nan=False)


def _template_values_match(actual, expected, *, rtol=1e-12, atol=1e-14):
    """Compare reread Excel values without treating harmless float serialization as corruption."""
    if isinstance(actual, Real) and isinstance(expected, Real) and not isinstance(actual, bool) and not isinstance(expected, bool):
        return bool(np.isclose(float(actual), float(expected), rtol=rtol, atol=atol, equal_nan=False))
    return actual == expected


def fill_excel_template(source, destination, sheet_name, values, *, verify_cells=True, verify_rtol=1e-12, verify_atol=1e-14):
    """Copy an official template, fill explicit cell addresses, and preserve its structure."""
    source, destination = Path(source), Path(destination)
    if source.resolve() == destination.resolve():
        raise ValueError("destination must not overwrite the input template")
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    workbook = load_workbook(destination)
    worksheet = workbook[sheet_name]
    for address, value in values.items():
        worksheet[address] = value
    workbook.save(destination)
    if verify_cells:
        checked = load_workbook(destination, data_only=False)[sheet_name]
        mismatches = [
            address
            for address, value in values.items()
            if not _template_values_match(checked[address].value, value, rtol=verify_rtol, atol=verify_atol)
        ]
        if mismatches:
            raise RuntimeError(f"template cell verification failed: {mismatches}")
    return destination
