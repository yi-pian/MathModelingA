from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from core.export import export_origin_table, fill_excel_template, verify_excel, write_excel_checked


def test_multisheet_excel_write_and_verify(tmp_path):
    path = tmp_path / "result.xlsx"
    data = {"Q1": pd.DataFrame({"time_s": [0.0, 1.0], "x_m": [1.0, 2.0]}), "Summary": pd.DataFrame({"metric": ["max"], "value": [2.0]})}
    verification = write_excel_checked(path, data, column_orders={"Q1": ["time_s", "x_m"]}, decimals=4)
    assert verification["valid"] and verification["sheets"]["Q1"]["rows"] == 2
    assert verify_excel(path)["valid"]


def test_column_mapping_becomes_one_results_sheet(tmp_path):
    path = tmp_path / "single.xlsx"
    verification = write_excel_checked(path, {"time_s": [0, 1], "value": [2, 3]})
    assert verification["valid"] and list(pd.read_excel(path).columns) == ["time_s", "value"]


def test_origin_x_first_and_template_preservation(tmp_path):
    origin = tmp_path / "trajectory.xlsx"
    export_origin_table(origin, {"y_m": [2, 3], "time_s": [0, 1]}, x_column="time_s", metadata={"purpose": "trajectory"})
    frame = pd.read_excel(origin, sheet_name="Data")
    assert list(frame.columns) == ["time_s", "y_m"]
    source = tmp_path / "template.xlsx"; destination = tmp_path / "filled.xlsx"
    workbook = Workbook(); sheet = workbook.active; sheet.title = "Result"; sheet["A1"] = "fixed header"; workbook.save(source)
    fill_excel_template(source, destination, "Result", {"B2": 12.5})
    checked = load_workbook(destination)["Result"]
    assert checked["A1"].value == "fixed header" and checked["B2"].value == 12.5


def test_template_float_roundtrip_uses_numerical_tolerance(tmp_path):
    source = tmp_path / "template.xlsx"
    destination = tmp_path / "filled.xlsx"
    workbook = Workbook()
    workbook.active.title = "Result"
    workbook.save(source)
    # This value is serialized by XLSX/openpyxl as the adjacent binary float.
    value = 0.15968408615271582
    fill_excel_template(source, destination, "Result", {"A1": value})
    actual = load_workbook(destination, data_only=False)["Result"]["A1"].value
    assert actual != value
    assert abs(actual - value) < 1e-15


def test_excel_export_rejects_infinity(tmp_path):
    with pytest.raises(ValueError, match="contains Inf"):
        write_excel_checked(tmp_path / "inf.xlsx", pd.DataFrame({"value": [1.0, np.inf]}))


def test_excel_verification_reports_infinity(tmp_path):
    path = tmp_path / "existing_inf.xlsx"
    pd.DataFrame({"value": [1.0, np.inf]}).to_excel(path, index=False)
    checked = verify_excel(path)
    assert not checked["valid"]
    assert checked["sheets"]["Sheet1"]["inf_count"] == 1
