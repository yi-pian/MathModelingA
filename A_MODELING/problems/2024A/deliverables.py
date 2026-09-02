"""Official-template and reporting helpers specific to CUMCM 2024A."""

from __future__ import annotations

from pathlib import Path
import shutil
import sys

import numpy as np
import pandas as pd
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
PROJECT_ROOT = next(parent for parent in HERE.parents if (parent / "core").is_dir())
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core.export import fill_excel_template, verify_excel

from common import NODE_LABELS, N_HANDLES, SELECTED_NODE_INDICES


OFFICIAL_DIR = PROJECT_ROOT / "data" / "2024A" / "official" / "extracted" / "A题" / "附件"
RESULT_DIR = PROJECT_ROOT / "results" / "2024A"


def fill_time_series_template(source: Path, destination: Path, times, positions, speeds):
    """Fill result1/result4 while preserving their two-sheet official structure."""
    source, destination = Path(source), Path(destination)
    times = np.asarray(times, float)
    positions, speeds = np.asarray(positions, float), np.asarray(speeds, float)
    if positions.shape != (len(times), N_HANDLES, 2) or speeds.shape != (len(times), N_HANDLES):
        raise ValueError("unexpected time-series array shape")
    if not np.all(np.isfinite(positions)) or not np.all(np.isfinite(speeds)):
        raise ValueError("official output contains NaN or Inf")
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    workbook = load_workbook(destination)
    position_sheet, speed_sheet = workbook["位置"], workbook["速度"]
    if position_sheet.max_column != len(times) + 1 or speed_sheet.max_column != len(times) + 1:
        raise ValueError("official template time-column count does not match")
    for time_index in range(len(times)):
        column = time_index + 2
        for node in range(N_HANDLES):
            x_cell = position_sheet.cell(row=2 + 2 * node, column=column)
            y_cell = position_sheet.cell(row=3 + 2 * node, column=column)
            speed_cell = speed_sheet.cell(row=2 + node, column=column)
            x_cell.value = round(float(positions[time_index, node, 0]), 6)
            y_cell.value = round(float(positions[time_index, node, 1]), 6)
            speed_cell.value = round(float(speeds[time_index, node]), 6)
            x_cell.number_format = y_cell.number_format = speed_cell.number_format = "0.000000"
    workbook.save(destination)
    return verify_time_series_workbook(source, destination, times)


def fill_result2_template(destination: Path, positions, speeds):
    positions, speeds = np.asarray(positions, float), np.asarray(speeds, float)
    if positions.shape != (N_HANDLES, 2) or speeds.shape != (N_HANDLES,) or not np.all(np.isfinite(positions)) or not np.all(np.isfinite(speeds)):
        raise ValueError("unexpected result2 arrays")
    values = {}
    for node in range(N_HANDLES):
        row = node + 2
        values[f"B{row}"] = round(float(positions[node, 0]), 6)
        values[f"C{row}"] = round(float(positions[node, 1]), 6)
        values[f"D{row}"] = round(float(speeds[node]), 6)
    destination = fill_excel_template(OFFICIAL_DIR / "result2.xlsx", destination, "Sheet1", values)
    workbook = load_workbook(destination)
    for row in range(2, N_HANDLES + 2):
        for column in range(2, 5):
            workbook["Sheet1"].cell(row=row, column=column).number_format = "0.000000"
    workbook.save(destination)
    return verify_result2_workbook(OFFICIAL_DIR / "result2.xlsx", destination)


def verify_time_series_workbook(source, destination, times):
    source_book = load_workbook(source, data_only=False)
    output_book = load_workbook(destination, data_only=False)
    errors = []
    if output_book.sheetnames != source_book.sheetnames:
        errors.append("sheet names/order changed")
    for name, expected_rows in (("位置", 449), ("速度", 225)):
        source_sheet, output_sheet = source_book[name], output_book[name]
        if (output_sheet.max_row, output_sheet.max_column) != (expected_rows, len(times) + 1):
            errors.append(f"{name}: wrong shape")
        if [output_sheet.cell(1, col).value for col in range(1, output_sheet.max_column + 1)] != [source_sheet.cell(1, col).value for col in range(1, source_sheet.max_column + 1)]:
            errors.append(f"{name}: time header changed")
        if [output_sheet.cell(row, 1).value for row in range(1, output_sheet.max_row + 1)] != [source_sheet.cell(row, 1).value for row in range(1, source_sheet.max_row + 1)]:
            errors.append(f"{name}: node labels/order changed")
        values = np.array([[output_sheet.cell(row, col).value for col in range(2, output_sheet.max_column + 1)] for row in range(2, output_sheet.max_row + 1)], dtype=float)
        if not np.all(np.isfinite(values)):
            errors.append(f"{name}: NaN or Inf")
    generic = verify_excel(destination, allow_nan=False)
    if not generic["valid"]:
        errors.extend(generic["errors"])
    return {"valid": not errors, "errors": errors, "path": str(destination), "times": [float(times[0]), float(times[-1])], "shape_position": [449, len(times) + 1], "shape_speed": [225, len(times) + 1]}


def verify_result2_workbook(source, destination):
    source_book, output_book = load_workbook(source), load_workbook(destination)
    source_sheet, output_sheet = source_book["Sheet1"], output_book["Sheet1"]
    errors = []
    if output_book.sheetnames != ["Sheet1"] or (output_sheet.max_row, output_sheet.max_column) != (225, 4):
        errors.append("wrong result2 workbook structure")
    if [output_sheet.cell(row, 1).value for row in range(1, 226)] != [source_sheet.cell(row, 1).value for row in range(1, 226)]:
        errors.append("node labels/order changed")
    if [output_sheet.cell(1, col).value for col in range(1, 5)] != [source_sheet.cell(1, col).value for col in range(1, 5)]:
        errors.append("column headers changed")
    values = np.array([[output_sheet.cell(row, col).value for col in range(2, 5)] for row in range(2, 226)], dtype=float)
    if not np.all(np.isfinite(values)):
        errors.append("NaN or Inf")
    return {"valid": not errors, "errors": errors, "path": str(destination), "shape": [225, 4], "labels": NODE_LABELS}


def selected_position_table(times, positions):
    rows = []
    for node in SELECTED_NODE_INDICES:
        row_x = {"node": NODE_LABELS[node], "coordinate": "x (m)"}
        row_y = {"node": NODE_LABELS[node], "coordinate": "y (m)"}
        for time_index, time in enumerate(times):
            row_x[f"{time:g} s"] = float(positions[time_index, node, 0])
            row_y[f"{time:g} s"] = float(positions[time_index, node, 1])
        rows.extend((row_x, row_y))
    return pd.DataFrame(rows)


def selected_speed_table(times, speeds):
    rows = []
    for node in SELECTED_NODE_INDICES:
        row = {"node": NODE_LABELS[node], "unit": "m/s"}
        for time_index, time in enumerate(times):
            row[f"{time:g} s"] = float(speeds[time_index, node])
        rows.append(row)
    return pd.DataFrame(rows)
