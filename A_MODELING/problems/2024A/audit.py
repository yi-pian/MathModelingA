"""Independent read-only reverse audit of the finished 2024A computation."""

from __future__ import annotations

from pathlib import Path
import json
import re
import sys

import numpy as np
import pandas as pd
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
ROOT = next(parent for parent in HERE.parents if (parent / "core").is_dir())
sys.path.insert(0, str(HERE)); sys.path.insert(0, str(ROOT))

from common import (
    ArchimedeanSpiral, HEAD_HANDLE_DISTANCE_M, BODY_HANDLE_DISTANCE_M,
    LINK_LENGTHS_M, TURN_RADIUS_M, TurnaroundPath, build_path_chain,
    build_spiral_chain, minimum_bench_clearance, pair_clearance,
)
from deliverables import OFFICIAL_DIR, RESULT_DIR


def load_metrics(question):
    return json.loads((RESULT_DIR / f"q{question}_metrics.json").read_text(encoding="utf-8"))


def workbook_matrix(path, sheet_name):
    sheet = load_workbook(path, data_only=True, read_only=True)[sheet_name]
    # In read-only mode repeated ``cell()`` calls rescan worksheet XML.  Stream
    # rows once so this independent audit remains O(number of cells).
    values = ([cell for cell in row[1:]] for row in sheet.iter_rows(min_row=2, values_only=True))
    return np.asarray(list(values), dtype=float)


def audit_template(source, output, expected_headers):
    source_book = load_workbook(source, data_only=False)
    output_book = load_workbook(output, data_only=False)
    errors = []
    if output_book.sheetnames != source_book.sheetnames:
        errors.append("sheet names/order changed")
    for name in source_book.sheetnames:
        source_sheet, output_sheet = source_book[name], output_book[name]
        if (source_sheet.max_row, source_sheet.max_column) != (output_sheet.max_row, output_sheet.max_column):
            errors.append(f"{name}: dimensions changed")
        headers = [output_sheet.cell(1, column).value for column in range(2, output_sheet.max_column + 1)]
        if headers != expected_headers:
            errors.append(f"{name}: time headers/order wrong")
        source_labels = [source_sheet.cell(row, 1).value for row in range(1, source_sheet.max_row + 1)]
        output_labels = [output_sheet.cell(row, 1).value for row in range(1, output_sheet.max_row + 1)]
        if output_labels != source_labels:
            errors.append(f"{name}: node labels/order changed")
        for row in range(1, source_sheet.max_row + 1):
            for column in range(1, source_sheet.max_column + 1):
                source_cell, output_cell = source_sheet.cell(row, column), output_sheet.cell(row, column)
                # Filling the official blank data region deliberately changes only
                # numFmtId to enforce six displayed decimals.  Every visual style
                # component must otherwise remain identical.
                source_style = source_cell._style
                output_style = output_cell._style
                visual_source = (source_style.fontId, source_style.fillId, source_style.borderId, source_style.alignmentId, source_style.protectionId, source_style.xfId)
                visual_output = (output_style.fontId, output_style.fillId, output_style.borderId, output_style.alignmentId, output_style.protectionId, output_style.xfId)
                if visual_source != visual_output:
                    errors.append(f"{name}: visual style changed at {row},{column}")
                    break
                if row >= 2 and column >= 2 and output_cell.number_format != "0.000000":
                    errors.append(f"{name}: data format is not six decimals at {row},{column}")
                    break
                if (row == 1 or column == 1) and source_cell.number_format != output_cell.number_format:
                    errors.append(f"{name}: header/label number format changed at {row},{column}")
                    break
            if errors and (errors[-1].startswith(f"{name}: visual") or errors[-1].startswith(f"{name}: data") or errors[-1].startswith(f"{name}: header")):
                break
    formula_count = sum(
        1 for sheet in output_book.worksheets for row in sheet.iter_rows() for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    return {"valid": not errors, "errors": errors, "formula_count": formula_count, "sheets": output_book.sheetnames}


def nearest_root_audit_spiral(state, pitch_m):
    spiral = ArchimedeanSpiral(pitch_m)
    failures = 0
    for index, target in enumerate(LINK_LENGTHS_M, start=1):
        left, right = state.coordinates[index - 1], state.coordinates[index]
        previous = state.positions[index - 1]
        probes = np.linspace(left, right, 21)[1:-1]
        if np.any([np.linalg.norm(spiral.point(theta) - previous) >= target + 2e-9 for theta in probes]):
            failures += 1
    return failures


def nearest_root_audit_path(path, state):
    failures = 0
    for index, target in enumerate(LINK_LENGTHS_M, start=1):
        earlier, later = state.coordinates[index], state.coordinates[index - 1]
        previous = state.positions[index - 1]
        probes = np.linspace(later, earlier, 21)[1:-1]
        if np.any([np.linalg.norm(path.point(value) - previous) >= target + 2e-9 for value in probes]):
            failures += 1
    return failures


def main():
    metrics = {question: load_metrics(question) for question in range(1, 6)}
    checks: list[tuple[str, bool, str]] = []
    add = lambda name, passed, detail="": checks.append((name, bool(passed), str(detail)))
    add("Q1-Q5 status", all(metrics[q]["status"] == "PASS" for q in metrics))

    q1_raw = np.load(RESULT_DIR / "raw" / "q1.npz")
    q4_raw = np.load(RESULT_DIR / "raw" / "q4.npz")
    q1_excel = audit_template(OFFICIAL_DIR / "result1.xlsx", RESULT_DIR / "result1.xlsx", [f"{time} s" for time in range(301)])
    q4_excel = audit_template(OFFICIAL_DIR / "result4.xlsx", RESULT_DIR / "result4.xlsx", [f"{time} s" for time in range(-100, 101)])
    q1_position_expected = np.empty((448, 301)); q1_position_expected[0::2] = q1_raw["positions"][:, :, 0].T; q1_position_expected[1::2] = q1_raw["positions"][:, :, 1].T
    q4_position_expected = np.empty((448, 201)); q4_position_expected[0::2] = q4_raw["positions"][:, :, 0].T; q4_position_expected[1::2] = q4_raw["positions"][:, :, 1].T
    q1_position_error = float(np.max(np.abs(workbook_matrix(RESULT_DIR / "result1.xlsx", "位置") - np.round(q1_position_expected, 6))))
    q1_speed_error = float(np.max(np.abs(workbook_matrix(RESULT_DIR / "result1.xlsx", "速度") - np.round(q1_raw["speeds"].T, 6))))
    q4_position_error = float(np.max(np.abs(workbook_matrix(RESULT_DIR / "result4.xlsx", "位置") - np.round(q4_position_expected, 6))))
    q4_speed_error = float(np.max(np.abs(workbook_matrix(RESULT_DIR / "result4.xlsx", "速度") - np.round(q4_raw["speeds"].T, 6))))
    add("result1 template/header/style", q1_excel["valid"], q1_excel["errors"])
    add("result1 values equal final arrays", max(q1_position_error, q1_speed_error) == 0, f"max={max(q1_position_error, q1_speed_error):.3e}")
    add("result4 template/header/style", q4_excel["valid"], q4_excel["errors"])
    add("result4 values equal final arrays", max(q4_position_error, q4_speed_error) == 0, f"max={max(q4_position_error, q4_speed_error):.3e}")

    q2_spiral = ArchimedeanSpiral(0.55)
    q2_theta = q2_spiral.theta_from_arc(q2_spiral.arc_primitive(32.0 * np.pi) - metrics[2]["event_time_s"], upper_hint=32.0 * np.pi)[0]
    q2_state = build_spiral_chain(q2_theta)
    q2_matrix = workbook_matrix(RESULT_DIR / "result2.xlsx", "Sheet1")
    q2_expected = np.column_stack((q2_state.positions, q2_state.speeds))
    q2_value_error = float(np.max(np.abs(q2_matrix - np.round(q2_expected, 6))))
    q2_pair = tuple(metrics[2]["collision_pair_bench_indices"])
    add("result2 values equal critical state", q2_value_error == 0, f"max={q2_value_error:.3e}")
    add("Q2 physical-rectangle event", abs(pair_clearance(q2_state.positions, q2_pair)) < 1e-9)
    add("Q2 event direction", metrics[2]["global_clearance_before_m"] > 0 > metrics[2]["global_clearance_after_m"])
    add("Q2 nearest spiral roots", nearest_root_audit_spiral(q2_state, 0.55) == 0)

    q3_state = build_spiral_chain(metrics[3]["critical_theta_head_rad"], pitch_m=metrics[3]["critical_pitch_m"])
    add("Q3 critical physical contact", abs(minimum_bench_clearance(q3_state.positions).clearance) < 1e-8)
    add("Q3 feasibility direction", metrics[3]["clearance_below_m"] < 0 < metrics[3]["clearance_above_m"])
    add("Q3 full approach, not endpoint only", metrics[3]["critical_head_radius_m"] > TURN_RADIUS_M and metrics[3]["outer_approach_min_clearance_m"] > 0)
    add("Q3 nearest spiral roots", nearest_root_audit_spiral(q3_state, metrics[3]["critical_pitch_m"]) == 0)

    path = TurnaroundPath()
    q5_state = build_path_chain(path, metrics[5]["critical_head_path_coordinate_m"], head_speed_m_s=metrics[5]["maximum_head_speed_m_s"])
    centers_distance = float(np.linalg.norm(path.center2 - path.center1))
    add("Q4 circle tangency", abs(centers_distance - path.radius1 - path.radius2) < 1e-10)
    add("Q4 radius ratio and boundary", abs(path.radius1 / path.radius2 - 2.0) < 1e-12 and path.max_turn_radius(4001) <= 4.5 + 1e-10)
    add("Q4 nearest path roots", nearest_root_audit_path(path, build_path_chain(path, 14.0)) == 0)
    add("Q5 maximize direction", metrics[5]["maximum_unit_head_speed_ratio"] > max(item["ratio"] for item in metrics[5]["refined_candidates"][1:]))
    add("Q5 speed cap active", abs(np.max(q5_state.speeds) - 2.0) < 1e-10)
    audited_limiting_nodes = np.flatnonzero(np.max(q5_state.speeds) - q5_state.speeds <= 1e-10).tolist()
    add("Q5 limiting-handle set", audited_limiting_nodes == metrics[5]["limiting_handle_indices"], audited_limiting_nodes)
    add("Q5 perturbation direction", metrics[5]["max_speed_below_limit_m_s"] < 2.0 < metrics[5]["max_speed_above_limit_m_s"])
    add("Q5 nearest path roots", nearest_root_audit_path(path, q5_state) == 0)

    add("Handle distances from dimensions", abs(HEAD_HANDLE_DISTANCE_M - 2.86) < 1e-12 and abs(BODY_HANDLE_DISTANCE_M - 1.65) < 1e-12)
    source_text = "\n".join(path.read_text(encoding="utf-8") for path in HERE.glob("*.py"))
    degree_calls = re.findall(r"np\.(?:deg2rad|rad2deg|degrees|radians)\s*\(", source_text)
    add("No degree/radian conversion in solver", len(degree_calls) == 0, degree_calls)
    add("Finite raw arrays", all(np.all(np.isfinite(array)) for array in (q1_raw["positions"], q1_raw["speeds"], q4_raw["positions"], q4_raw["speeds"])))
    add("Monotone unique Q1/Q4 time", np.all(np.diff(q1_raw["times"]) > 0) and np.all(np.diff(q4_raw["times"]) > 0))

    expected_figures = [
        "q1_shapes", "q1_speeds", "q2_critical_collision", "q2_event_clearance",
        "q3_pitch_clearance", "q3_critical_contact", "q4_path_shapes_inbound",
        "q4_path_shapes_outbound", "q4_speeds", "q5_speed_limit",
    ]
    missing_figures = [f"{name}.{suffix}" for name in expected_figures for suffix in ("png", "pdf", "svg") if not (RESULT_DIR / "figures" / f"{name}.{suffix}").exists()]
    add("Paper figures PNG/PDF/SVG", not missing_figures, missing_figures)
    origin_errors = []
    for workbook in (RESULT_DIR / "origin_data").glob("*.xlsx"):
        sheets = pd.read_excel(workbook, sheet_name=None)
        if not sheets or any(frame.empty or frame.isna().any().any() for frame in sheets.values()):
            origin_errors.append(workbook.name)
    add("Origin workbooks finite/nonempty", not origin_errors, origin_errors)
    add("Per-link error workbook", (RESULT_DIR / "chain_error_by_link.xlsx").exists())

    passed = all(item[1] for item in checks)
    audit_metrics = {
        "status": "PASS" if passed else "FAIL", "checks": [
            {"name": name, "passed": ok, "detail": detail} for name, ok, detail in checks
        ],
        "excel_value_max_errors": {
            "result1_position": q1_position_error, "result1_speed": q1_speed_error,
            "result2": q2_value_error, "result4_position": q4_position_error,
            "result4_speed": q4_speed_error,
        },
    }
    (RESULT_DIR / "audit_metrics.json").write_text(json.dumps(audit_metrics, ensure_ascii=False, indent=2), encoding="utf-8")
    lines = ["# 2024A 反向审查报告", "", f"结论：**{audit_metrics['status']}**。", ""]
    lines.extend(f"- {'PASS' if ok else 'FAIL'} — {name}" + (f"：{detail}" if detail else "") for name, ok, detail in checks)
    lines.extend(["", "审查以官方模板、最终原始数组和独立重算状态为依据；未使用公开完整解答或网上代码。", ""])
    (RESULT_DIR / "AUDIT_REPORT.md").write_text("\n".join(lines), encoding="utf-8")
    print("\n".join(lines))
    if not passed:
        raise RuntimeError("reverse audit failed")


if __name__ == "__main__": main()
