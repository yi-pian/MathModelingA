import sys
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parents[1]
ROOT = next(parent for parent in HERE.parents if (parent / "core").is_dir())
sys.path.insert(0, str(HERE)); sys.path.insert(0, str(ROOT))

from common import N_HANDLES
from deliverables import OFFICIAL_DIR, fill_result2_template, fill_time_series_template


def test_result2_template_copy_and_order(tmp_path):
    positions = np.column_stack((np.arange(N_HANDLES), -np.arange(N_HANDLES))).astype(float)
    speeds = np.linspace(1, 2, N_HANDLES)
    destination = tmp_path / "result2.xlsx"
    verification = fill_result2_template(destination, positions, speeds)
    assert verification["valid"]
    workbook = load_workbook(destination)
    assert workbook["Sheet1"]["B2"].value == 0
    assert workbook["Sheet1"]["B225"].value == N_HANDLES - 1


def test_result1_template_copy_and_mapping(tmp_path):
    times = np.arange(301, dtype=float)
    positions = np.zeros((301, N_HANDLES, 2)); speeds = np.ones((301, N_HANDLES))
    positions[:, :, 0] = np.arange(N_HANDLES)[None, :]
    positions[:, :, 1] = times[:, None]
    destination = tmp_path / "result1.xlsx"
    verification = fill_time_series_template(OFFICIAL_DIR / "result1.xlsx", destination, times, positions, speeds)
    assert verification["valid"]
    workbook = load_workbook(destination, data_only=True)
    assert workbook["位置"].cell(row=448, column=302).value == N_HANDLES - 1
    assert workbook["位置"].cell(row=449, column=302).value == 300
    assert workbook["速度"].cell(row=225, column=302).value == 1
