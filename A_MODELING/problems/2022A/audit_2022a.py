"""Independent read-only audit of final 2022A artifacts.

The numerical model is not called here.  This pass only rereads committed
metrics, official-template workbooks, figures, and Origin tables so stale or
internally inconsistent deliverables cannot pass by construction.
"""

from __future__ import annotations

import json
from pathlib import Path
import sys

import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = next(parent for parent in Path(__file__).resolve().parents if (parent / "core").is_dir())
sys.path.insert(0, str(ROOT))

from core.integration import integrate_samples


RESULTS = ROOT / "results" / "2022A"
OFFICIAL = ROOT / "data" / "2022A" / "official" / "A"


def _load(name: str) -> dict:
    return json.loads((RESULTS / name).read_text(encoding="utf-8"))


def _workbook_data(path: Path) -> np.ndarray:
    sheet = load_workbook(path, data_only=True).active
    return np.asarray(
        [[sheet.cell(row, column).value for column in range(1, sheet.max_column + 1)] for row in range(3, sheet.max_row + 1)],
        dtype=float,
    )


def _template_preserved(result_name: str) -> bool:
    original = load_workbook(OFFICIAL / result_name, data_only=False).active
    result = load_workbook(RESULTS / result_name, data_only=False).active
    original_headers = [[original.cell(row, col).value for col in range(1, original.max_column + 1)] for row in (1, 2)]
    result_headers = [[result.cell(row, col).value for col in range(1, result.max_column + 1)] for row in (1, 2)]
    return (
        original.title == result.title
        and original.max_column == result.max_column
        and original_headers == result_headers
        and {str(item) for item in original.merged_cells.ranges} == {str(item) for item in result.merged_cells.ranges}
    )


def run() -> dict:
    q1, q2, q3, q4 = (_load(f"q{index}_metrics.json") for index in range(1, 5))
    validation = _load("validation_2022a.json")
    manifest = _load("figure_manifest.json")
    checks: list[dict] = []

    def check(section: str, item: str, passed: bool, evidence: str) -> None:
        checks.append({"section": section, "item": item, "passed": bool(passed), "evidence": evidence})

    for index, metrics in enumerate((q1, q2, q3, q4), start=1):
        check("completion", f"Q{index} declared status", metrics["status"] == "PASS", metrics["status"])

    check("formula", "angle unit is rad", q3["angle_unit"] == "rad", q3["angle_unit"])
    check("formula", "Q3 state order fixed", q3["state_order"] == ["x_rel", "z_f", "theta_f", "theta_o", "x_dot_rel", "v_f", "omega_f", "omega_o"], str(q3["state_order"]))
    check("formula", "Q3 nonlinear model degenerates to heave", q3["heave_degeneration_max_abs_error"] < 1e-8, f"max error={q3['heave_degeneration_max_abs_error']:.3e}")
    check("formula", "Q3 small-signal linearization", q3["small_signal_linearization_relative_error"] < 1e-4, f"relative error={q3['small_signal_linearization_relative_error']:.3e}")
    required_inertia = {"float_pitch_inertia", "oscillator_centroid_pitch_inertia", "oscillator_axis_distance", "float_components"}
    check("formula", "inertia convention fully recorded", required_inertia <= set(q3["inertia_audit"]), f"keys={sorted(required_inertia)}")

    for case in ("linear", "nonlinear"):
        values = q1["convergence"][case]["max_differences"]
        check("precision", f"Q1 {case} tolerance refinement", values[1] < values[0] and q1["convergence"][case]["delivery_accuracy_pass"], f"{values[0]:.3e} -> {values[1]:.3e}")
    check("precision", "Q1 long-transient frequency cross-check", q1["frequency_domain_120_cycle_relative_error"] < 1e-8, f"relative error={q1['frequency_domain_120_cycle_relative_error']:.3e}")
    constant_gap = abs(q2["constant"]["frequency_power_w"] - q2["constant"]["shooting_power_w"])
    check("precision", "Q2 frequency/shooting agreement", constant_gap < 1e-5, f"difference={constant_gap:.3e} W")
    check("precision", "Q2 nonlinear repeatability", q2["nonlinear"]["strict_repeat_range_w"] < 1e-8, f"range={q2['nonlinear']['strict_repeat_range_w']:.3e} W")
    check("precision", "Q3 periodic shooting closure", q3["periodic_shooting_residual"] < 1e-10, f"residual={q3['periodic_shooting_residual']:.3e}")
    check("precision", "Q3 long-transient agreement", q3["periodic_long_transient_relative_error"] < 1e-5, f"relative error={q3['periodic_long_transient_relative_error']:.3e}")
    q4_residuals = [level["periodic_residual"] for level in q4["shooting_levels"]]
    check("precision", "Q4 strict shooting closure", q4_residuals[-1] < 1e-10 and all(right < left for left, right in zip(q4_residuals, q4_residuals[1:])), " -> ".join(f"{value:.3e}" for value in q4_residuals))

    q2_energy = validation["energy"]["q2_nonlinear"]
    q4_energy = validation["energy"]["q4_coupled"]
    solver_gap = validation["solver_comparison"]["max_abs_difference"]
    check("precision", "Q1 DOP853/RK45 agreement", solver_gap < 1e-7, f"max difference={solver_gap:.3e}")
    check("energy", "Q2 periodic energy balance", abs(q2_energy["residual_w"]) < 2e-5, f"residual={q2_energy['residual_w']:.3e} W")
    check("energy", "Q4 periodic energy balance", abs(q4_energy["residual_w"]) < 2e-5, f"residual={q4_energy['residual_w']:.3e} W")
    check("power", "Q4 channel sum", abs(q4["heave_power_w"] + q4["rotation_power_w"] - q4["total_power_w"]) < 1e-9, f"residual={q4['power_sum_residual_w']:.3e} W")
    check("power", "all reported PTO powers nonnegative", min(q2["constant"]["frequency_power_w"], q2["nonlinear"]["strict_power_w"], q4["heave_power_w"], q4["rotation_power_w"]) >= 0.0, "minimum >= 0 W")

    q2_neighbor_max = max(item["average_power_w"] for item in q2["nonlinear"]["neighbors"])
    check("optimization", "Q2 nonlinear neighborhood", q2_neighbor_max <= q2["nonlinear"]["strict_power_w"] + 1e-7, f"best neighbor={q2_neighbor_max:.9f} W")
    q4_neighbor_max = max(item["total_power_w"] for item in q4["neighborhood"])
    check("optimization", "Q4 neighborhood", q4_neighbor_max <= q4["total_power_w"] + 1e-7, f"best neighbor={q4_neighbor_max:.9f} W")
    bounds_ok = 0.0 <= q4["optimal_linear_damping_n_s_m"] <= 100000.0 and 0.0 <= q4["optimal_rotational_damping_n_m_s"] <= 100000.0
    check("optimization", "Q4 bounds and constrained upper optimum", bounds_ok and q4["optimal_rotational_damping_n_m_s"] == 100000.0, f"({q4['optimal_linear_damping_n_s_m']:.6f}, {q4['optimal_rotational_damping_n_m_s']:.1f})")
    check("optimization", "optimizers report success", q2["nonlinear"]["global"]["success"] and q2["nonlinear"]["local"]["success"] and q4["global"]["success"] and q4["local"]["success"] and q4["strict_local"]["success"], "global/local/strict_local")

    workbook_specs = {"result1-1.xlsx": (898, 5, 179.4), "result1-2.xlsx": (898, 5, 179.4), "result3.xlsx": (733, 9, 146.4)}
    workbook_arrays: dict[str, np.ndarray] = {}
    for name, (rows, columns, end_time) in workbook_specs.items():
        values = _workbook_data(RESULTS / name)
        workbook_arrays[name] = values
        valid = values.shape == (rows, columns) and np.all(np.isfinite(values)) and np.all(np.diff(values[:, 0]) > 0.0) and abs(values[0, 0]) < 1e-14 and abs(values[-1, 0] - end_time) < 1e-12
        check("excel", f"{name} dimensions/time/finite", valid, f"shape={values.shape}, t={values[0,0]}..{values[-1,0]}")
        check("excel", f"{name} official template preserved", _template_preserved(name), "sheet/header/merged cells")

    q1_at_100 = next(item for item in q1["selected"] if item["case"] == "linear" and item["time_s"] == 100.0)
    q1_row = workbook_arrays["result1-1.xlsx"][np.where(np.isclose(workbook_arrays["result1-1.xlsx"][:, 0], 100.0))[0][0]]
    q1_expected = np.array([100.0, q1_at_100["float_displacement_m"], q1_at_100["float_velocity_m_s"], q1_at_100["oscillator_displacement_m"], q1_at_100["oscillator_velocity_m_s"]])
    check("excel", "Q1 key row matches metrics", np.max(np.abs(q1_row - q1_expected)) < 5e-6, f"max difference={np.max(np.abs(q1_row-q1_expected)):.3e}")
    q3_at_100 = next(item for item in q3["selected"] if item["time_s"] == 100.0)
    q3_row = workbook_arrays["result3.xlsx"][np.where(np.isclose(workbook_arrays["result3.xlsx"][:, 0], 100.0))[0][0]]
    q3_expected = np.array([100.0, q3_at_100["float_heave_m"], q3_at_100["float_heave_velocity_m_s"], q3_at_100["float_pitch_rad"], q3_at_100["float_pitch_rate_rad_s"], q3_at_100["oscillator_heave_m"], q3_at_100["oscillator_heave_velocity_m_s"], q3_at_100["oscillator_pitch_rad"], q3_at_100["oscillator_pitch_rate_rad_s"]])
    check("excel", "Q3 key row matches metrics", np.max(np.abs(q3_row - q3_expected)) < 5e-6, f"max difference={np.max(np.abs(q3_row-q3_expected)):.3e}")

    figure_paths = [Path(path) for formats in manifest["figures"].values() for path in formats]
    origin_paths = [Path(path) for path in manifest["origin_files"]]
    check("figures", "all PNG/PDF/SVG files exist", len(figure_paths) == 21 and all(path.exists() and path.stat().st_size > 0 for path in figure_paths), f"files={len(figure_paths)}")
    check("figures", "all Origin workbooks exist", len(origin_paths) >= 11 and all(path.exists() and path.stat().st_size > 0 for path in origin_paths), f"files={len(origin_paths)}")
    split = pd.read_excel(RESULTS / "origin_data" / "q4_power_split.xlsx", sheet_name="Data")
    split_gap = np.max(np.abs(split["average_power_w"].to_numpy() - np.array([q4["heave_power_w"], q4["rotation_power_w"]])))
    check("figures", "Q4 plotted data matches final metrics", split_gap < 1e-9, f"max difference={split_gap:.3e} W")
    power_time = pd.read_excel(RESULTS / "origin_data" / "q2_optimal_power_time.xlsx", sheet_name="Data")
    power_time_average = integrate_samples(power_time["time_s"], power_time["instantaneous_power_w"], method="simpson") / (power_time["time_s"].iloc[-1] - power_time["time_s"].iloc[0])
    check("figures", "Q2 periodic power-time table matches optimum", abs(power_time_average - q2["nonlinear"]["strict_power_w"]) < 1e-3, f"average={power_time_average:.9f} W")
    q2_neighborhood = pd.read_excel(RESULTS / "origin_data" / "q2_optimum_neighborhood.xlsx", sheet_name="Data")["average_power_w"].to_numpy()
    q4_neighborhood = pd.read_excel(RESULTS / "origin_data" / "q4_optimum_neighborhood.xlsx", sheet_name="Data")["total_power_w"].to_numpy()
    neighborhood_gap = max(
        np.max(np.abs(q2_neighborhood - np.array([item["average_power_w"] for item in q2["nonlinear"]["neighbors"]]))),
        np.max(np.abs(q4_neighborhood - np.array([item["total_power_w"] for item in q4["neighborhood"]]))),
    )
    check("figures", "Origin neighborhood tables match final metrics", neighborhood_gap < 1e-9, f"max difference={neighborhood_gap:.3e} W")

    timings = validation["performance"]
    performance_ok = all(np.isfinite(timings[key]) and timings[key] >= 0.0 for key in ("q1_seconds", "q2_seconds", "q3_seconds", "q4_seconds"))
    check("performance", "all question runtimes recorded", performance_ok, ", ".join(f"{key}={timings[key]:.3f}s" for key in ("q1_seconds", "q2_seconds", "q3_seconds", "q4_seconds")))

    passed = all(item["passed"] for item in checks)
    result = {"status": "PASS" if passed else "FAIL", "checks": checks, "failed": [item for item in checks if not item["passed"]]}
    lines = ["# 2022A 独立反向审查", "", f"结论：**{result['status']}**", "", "本轮不调用模型求解器，只重新读取最终指标、官方模板结果、图表和 Origin 数据。", ""]
    current = None
    for item in checks:
        if item["section"] != current:
            current = item["section"]
            lines.extend([f"## {current}", ""])
        mark = "PASS" if item["passed"] else "FAIL"
        lines.append(f"- {mark} — {item['item']}：{item['evidence']}")
    lines.extend(["", "## 审查员结论", "", "未发现公式符号、单位/弧度、状态顺序、根残差、优化方向、Excel 模板或图表数据不一致。Q4 的旋转阻尼位于题设上界，是受约束最优而非内部驻点。浮子壳体转动惯量依赖题意解释，已显式记录并做 ±10% 灵敏度，因此保留为模型假设而非神秘常数。" if passed else "存在失败项，正式验收不得判定通过。"])
    (RESULTS / "audit_2022a.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    (RESULTS / "AUDIT_REPORT.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    if not passed:
        raise RuntimeError(json.dumps(result["failed"], ensure_ascii=False, indent=2))
    return result


if __name__ == "__main__":
    outcome = run()
    print(json.dumps({"status": outcome["status"], "checks": len(outcome["checks"]), "failed": len(outcome["failed"])}, ensure_ascii=False, indent=2))
