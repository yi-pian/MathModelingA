"""Official-template Excel writing and strict reread checks for 2022A."""

from __future__ import annotations

from copy import copy
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from core.export import fill_excel_template


ROOT = Path(__file__).resolve().parents[2]
OFFICIAL = ROOT / "data" / "2022A" / "official" / "A"
RESULTS = ROOT / "results" / "2022A"


def _address_values(data: np.ndarray) -> dict[str, float]:
    values = np.asarray(data, float)
    if values.ndim != 2 or not np.all(np.isfinite(values)):
        raise ValueError("Excel data must be a finite two-dimensional array")
    return {
        f"{chr(64 + column)}{row + 3}": float(values[row, column - 1])
        for row in range(values.shape[0])
        for column in range(1, values.shape[1] + 1)
    }


def write_official_result(template_name: str, data: np.ndarray, *, expected_columns: int) -> dict:
    values = np.asarray(data, float)
    if values.ndim != 2 or values.shape[1] != expected_columns or not np.all(np.isfinite(values)):
        raise ValueError("official result shape or values are invalid")
    source = OFFICIAL / template_name
    destination = RESULTS / template_name
    RESULTS.mkdir(parents=True, exist_ok=True)
    fill_excel_template(source, destination, "Sheet1", _address_values(values))

    workbook = load_workbook(destination)
    worksheet = workbook["Sheet1"]
    style_source_row = 3
    for row in range(3, values.shape[0] + 3):
        worksheet.row_dimensions[row].height = worksheet.row_dimensions[style_source_row].height
        for column in range(1, expected_columns + 1):
            source_cell = worksheet.cell(style_source_row, column)
            cell = worksheet.cell(row, column)
            if row > 12:
                cell._style = copy(source_cell._style)
                cell.font = copy(source_cell.font)
                cell.fill = copy(source_cell.fill)
                cell.border = copy(source_cell.border)
                cell.alignment = copy(source_cell.alignment)
                cell.protection = copy(source_cell.protection)
            cell.number_format = "0.0" if column == 1 else "0.000000"
    workbook.save(destination)
    return verify_official_result(source, destination, values)


def verify_official_result(source: Path, destination: Path, expected: np.ndarray) -> dict:
    template = load_workbook(source, data_only=False)["Sheet1"]
    checked = load_workbook(destination, data_only=False)["Sheet1"]
    rows, columns = expected.shape
    actual = np.array(
        [[checked.cell(row + 3, column + 1).value for column in range(columns)] for row in range(rows)],
        dtype=float,
    )
    header_template = [[template.cell(row, column).value for column in range(1, columns + 1)] for row in (1, 2)]
    header_checked = [[checked.cell(row, column).value for column in range(1, columns + 1)] for row in (1, 2)]
    time = actual[:, 0]
    errors = []
    if checked.title != template.title:
        errors.append("sheet name changed")
    if set(map(str, checked.merged_cells.ranges)) != set(map(str, template.merged_cells.ranges)):
        errors.append("merged headers changed")
    if header_checked != header_template:
        errors.append("headers changed")
    if actual.shape != expected.shape:
        errors.append("shape changed")
    if not np.all(np.isfinite(actual)):
        errors.append("NaN or Inf found")
    if not np.all(np.diff(time) > 0.0):
        errors.append("time is not strictly increasing")
    if not np.allclose(actual, expected, rtol=0.0, atol=5e-13):
        errors.append("cell values changed after save")
    return {
        "valid": not errors,
        "errors": errors,
        "path": str(destination),
        "shape": list(actual.shape),
        "columns": columns,
        "time_start": float(time[0]),
        "time_end": float(time[-1]),
        "time_step_min": float(np.min(np.diff(time))),
        "time_step_max": float(np.max(np.diff(time))),
        "nan_count": int(np.isnan(actual).sum()),
        "inf_count": int(np.isinf(actual).sum()),
    }

