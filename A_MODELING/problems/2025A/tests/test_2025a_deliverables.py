import json
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from common import Strategy, obscuration_intervals, point_segment_distance_sq_many
from deliverables import RESULTS, SCHEMAS, verify_all


def test_cloud_behind_missile_is_not_on_sight_segment():
    missile = np.array([10.0, 0.0, 0.0])
    target = np.array([[0.0, 0.0, 0.0]])
    cloud_behind_missile = np.array([[12.0, 0.0, 0.0]])
    assert point_segment_distance_sq_many(cloud_behind_missile, missile, target)[0] == pytest.approx(4.0)


def test_cloud_behind_target_is_not_on_sight_segment():
    missile = np.array([10.0, 0.0, 0.0])
    target = np.array([[0.0, 0.0, 0.0]])
    cloud_behind_target = np.array([[-3.0, 0.0, 0.0]])
    assert point_segment_distance_sq_many(cloud_behind_target, missile, target)[0] == pytest.approx(9.0)


def test_q1_fast_and_final_are_consistent():
    strategy = Strategy("FY1", "M1", np.pi, 120.0, 1.5, 3.6)
    fast = obscuration_intervals(strategy, model="full", precision="FAST")
    final = obscuration_intervals(strategy, model="full", precision="FINAL")
    assert abs(fast.duration_s - final.duration_s) < 1e-3


def test_all_three_official_workbooks_pass_semantic_reread():
    report = verify_all(write=False)
    assert report["valid"], report


@pytest.mark.parametrize("name", ["result1.xlsx", "result2.xlsx", "result3.xlsx"])
def test_official_workbook_has_exact_required_shape_and_no_nonfinite(name):
    workbook = load_workbook(RESULTS / name, data_only=True)
    assert workbook.sheetnames == ["Sheet1"]
    sheet = workbook["Sheet1"]
    assert (sheet.max_row, sheet.max_column) == SCHEMAS[name]["shape"]
    numeric = [cell.value for row in sheet.iter_rows() for cell in row if isinstance(cell.value, (int, float))]
    assert np.all(np.isfinite(np.asarray(numeric, dtype=float)))


def test_q5_assignment_and_bomb_counts_are_legal():
    frame = pd.read_csv(RESULTS / "q5_strategy.csv")
    assert set(frame["missile"]) <= {"M1", "M2", "M3"}
    assert frame.groupby("uav")["bomb_no"].apply(list).to_dict() == {f"FY{i}": [1, 2, 3] for i in range(1, 6)}
    assert frame.groupby("uav")["heading_deg"].nunique().max() == 1
    assert frame.groupby("uav")["speed_m_s"].nunique().max() == 1


def test_audit_precision_and_excel_status_are_frozen_pass():
    payload = json.loads((Path(RESULTS) / "audit_result.json").read_text(encoding="utf-8"))
    assert payload["status"] == "PASS"
    assert payload["excel"]["valid"]
    assert max(item["absolute_difference_s"] for item in payload["precision"].values()) < 1e-3
