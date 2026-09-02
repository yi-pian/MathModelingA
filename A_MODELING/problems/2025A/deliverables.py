"""Strict official-template and semantic verification for CUMCM 2025A."""

from __future__ import annotations

import json
from pathlib import Path
import sys

import numpy as np
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
ROOT = next(parent for parent in HERE.parents if (parent / "core").is_dir())
sys.path.insert(0, str(HERE)); sys.path.insert(0, str(ROOT))

from core.units import degree_to_rad
from common import Strategy, burst_point, drop_point, obscuration_intervals, validate_drop_gaps
from problem_data import GRAVITY_M_S2, UAV_INITIAL_M

TEMPLATES = ROOT / "data" / "2025A" / "official" / "templates"
RESULTS = ROOT / "results" / "2025A"

SCHEMAS = {
    "result1.xlsx": {"shape": (6,10), "rows": range(2,5), "uav_col": None, "angle_col":1, "speed_col":2, "bomb_col":3, "drop_cols":(4,5,6), "burst_cols":(7,8,9), "duration_col":10, "missile_col":None},
    "result2.xlsx": {"shape": (6,10), "rows": range(2,5), "uav_col":1, "angle_col":2, "speed_col":3, "bomb_col":None, "drop_cols":(4,5,6), "burst_cols":(7,8,9), "duration_col":10, "missile_col":None},
    "result3.xlsx": {"shape": (18,12), "rows": range(2,17), "uav_col":1, "angle_col":2, "speed_col":3, "bomb_col":4, "drop_cols":(5,6,7), "burst_cols":(8,9,10), "duration_col":11, "missile_col":12},
}


def _column_dimensions(sheet):
    return {key:(value.width,value.hidden,value.outlineLevel) for key,value in sheet.column_dimensions.items()}


def _row_dimensions(sheet):
    return {key:(value.height,value.hidden,value.outlineLevel) for key,value in sheet.row_dimensions.items()}


def _strategy_from_row(name, sheet, row):
    schema=SCHEMAS[name]
    uav="FY1" if schema["uav_col"] is None else str(sheet.cell(row,schema["uav_col"]).value)
    missile="M1" if schema["missile_col"] is None else str(sheet.cell(row,schema["missile_col"]).value)
    bomb_no=1 if schema["bomb_col"] is None else int(sheet.cell(row,schema["bomb_col"]).value)
    angle_deg=float(sheet.cell(row,schema["angle_col"]).value);speed=float(sheet.cell(row,schema["speed_col"]).value)
    drop=np.array([sheet.cell(row,column).value for column in schema["drop_cols"]],dtype=float)
    burst=np.array([sheet.cell(row,column).value for column in schema["burst_cols"]],dtype=float)
    reported=float(sheet.cell(row,schema["duration_col"]).value)
    heading=float(degree_to_rad(angle_deg));direction=np.array([np.cos(heading),np.sin(heading)])
    delta_drop=drop[:2]-UAV_INITIAL_M[uav][:2]
    drop_time=float(np.dot(delta_drop,direction)/speed)
    cross_drop=float(np.linalg.norm(delta_drop-speed*drop_time*direction))
    delta_burst=burst[:2]-drop[:2]
    delay_xy=float(np.dot(delta_burst,direction)/speed)
    cross_burst=float(np.linalg.norm(delta_burst-speed*delay_xy*direction))
    vertical_loss=UAV_INITIAL_M[uav][2]-burst[2]
    delay_z=float(np.sqrt(max(0.0,2.0*vertical_loss/GRAVITY_M_S2)))
    strategy=Strategy(uav,missile,heading,speed,drop_time,delay_z,bomb_no)
    recomputed=obscuration_intervals(strategy,model="full",precision="FINAL")
    return {"strategy":strategy,"drop":drop,"burst":burst,"reported_duration_s":reported,"recomputed_duration_s":recomputed.duration_s,"drop_cross_track_m":cross_drop,"burst_cross_track_m":cross_burst,"delay_xy_s":delay_xy,"delay_z_s":delay_z,"drop_residual_m":float(np.linalg.norm(drop_point(strategy)-drop)),"burst_residual_m":float(np.linalg.norm(burst_point(strategy)-burst)),"duration_residual_s":abs(recomputed.duration_s-reported)}


def verify_workbook(name):
    schema=SCHEMAS[name];source_path=TEMPLATES/name;output_path=RESULTS/name
    source=load_workbook(source_path,data_only=False);output=load_workbook(output_path,data_only=False)
    errors=[]
    if output.sheetnames!=source.sheetnames:errors.append("sheet names/order changed")
    source_sheet=source["Sheet1"];sheet=output["Sheet1"]
    if (sheet.max_row,sheet.max_column)!=schema["shape"]:errors.append(f"shape {(sheet.max_row,sheet.max_column)} != {schema['shape']}")
    if list(sheet.merged_cells.ranges)!=list(source_sheet.merged_cells.ranges):errors.append("merged cells changed")
    if _column_dimensions(sheet)!=_column_dimensions(source_sheet):errors.append("column dimensions changed")
    if _row_dimensions(sheet)!=_row_dimensions(source_sheet):errors.append("row dimensions changed")
    if [sheet.cell(1,column).value for column in range(1,sheet.max_column+1)] != [source_sheet.cell(1,column).value for column in range(1,source_sheet.max_column+1)]:errors.append("headers changed")
    # ``style_id`` is workbook-local and style proxy objects compare by
    # identity across workbooks.  The immutable StyleArray records the actual
    # font/fill/border/alignment/format/protection references and is preserved
    # exactly because every result workbook starts from the official template.
    def style_key(cell,workbook):
        # An uninstantiated template cell has ``_style is None`` and inherits
        # the workbook's default cell style.  Once a value is written,
        # openpyxl may serialize that same default explicitly.
        style=cell._style if cell._style is not None else workbook._cell_styles[0]
        return tuple(style)
    if any(style_key(sheet.cell(row,column),output)!=style_key(source_sheet.cell(row,column),source) for row in range(1,sheet.max_row+1) for column in range(1,sheet.max_column+1)):errors.append("cell styles changed")
    note_row=schema["shape"][0]
    if [sheet.cell(note_row,column).value for column in range(1,sheet.max_column+1)] != [source_sheet.cell(note_row,column).value for column in range(1,source_sheet.max_column+1)]:errors.append("official note row changed")

    semantic=[]
    for row in schema["rows"]:
        required=[schema["angle_col"],schema["speed_col"],*schema["drop_cols"],*schema["burst_cols"],schema["duration_col"]]
        if schema["uav_col"]:required.append(schema["uav_col"])
        if schema["missile_col"]:required.append(schema["missile_col"])
        values=[sheet.cell(row,column).value for column in required]
        if any(value is None for value in values):errors.append(f"row {row}: blank required cell");continue
        numeric=[value for value in values if isinstance(value,(int,float))]
        if not np.all(np.isfinite(np.asarray(numeric,dtype=float))):errors.append(f"row {row}: NaN or Inf");continue
        try:item=_strategy_from_row(name,sheet,row)
        except Exception as exc:errors.append(f"row {row}: semantic reconstruction failed: {exc}");continue
        semantic.append(item)
        if item["drop_cross_track_m"]>2e-5:errors.append(f"row {row}: drop point not on UAV line")
        if item["burst_cross_track_m"]>2e-5:errors.append(f"row {row}: burst point horizontal mismatch")
        if abs(item["delay_xy_s"]-item["delay_z_s"])>2e-5:errors.append(f"row {row}: horizontal/vertical delay mismatch")
        if item["drop_residual_m"]>3e-5 or item["burst_residual_m"]>3e-5:errors.append(f"row {row}: kinematic coordinate residual too large")
        if item["duration_residual_s"]>2e-3:errors.append(f"row {row}: duration semantic residual {item['duration_residual_s']:.3e} s")

    strategies=[item["strategy"] for item in semantic]
    if name=="result1.xlsx" and [strategy.bomb_no for strategy in strategies]!=[1,2,3]:errors.append("result1 bomb order changed")
    if name=="result2.xlsx" and [strategy.uav for strategy in strategies]!=["FY1","FY2","FY3"]:errors.append("result2 UAV order changed")
    if name=="result3.xlsx":
        expected=[f"FY{index}" for index in range(1,6) for _ in range(3)]
        if [strategy.uav for strategy in strategies]!=expected:errors.append("result3 UAV order changed")
        if not validate_drop_gaps(strategies):errors.append("result3 same-UAV drop gap violated")
    return {"file":name,"valid":not errors,"errors":errors,"shape":[sheet.max_row,sheet.max_column],"rows_checked":len(semantic),"max_drop_residual_m":max((item["drop_residual_m"] for item in semantic),default=None),"max_burst_residual_m":max((item["burst_residual_m"] for item in semantic),default=None),"max_delay_consistency_s":max((abs(item["delay_xy_s"]-item["delay_z_s"]) for item in semantic),default=None),"max_duration_residual_s":max((item["duration_residual_s"] for item in semantic),default=None)}


def verify_all(write=True):
    reports={name:verify_workbook(name) for name in SCHEMAS}
    result={"valid":all(report["valid"] for report in reports.values()),"workbooks":reports}
    if write:
        (RESULTS/"excel_validation.json").write_text(json.dumps(result,ensure_ascii=False,indent=2),encoding="utf-8")
        lines=["2025A OFFICIAL EXCEL VALIDATION",f"overall={'PASS' if result['valid'] else 'FAIL'}"]
        for name,report in reports.items():lines.extend(["",f"[{name}] {'PASS' if report['valid'] else 'FAIL'}",f"shape={report['shape']}, rows={report['rows_checked']}",f"max_drop_residual_m={report['max_drop_residual_m']}",f"max_burst_residual_m={report['max_burst_residual_m']}",f"max_delay_consistency_s={report['max_delay_consistency_s']}",f"max_duration_residual_s={report['max_duration_residual_s']}",*report["errors"]])
        (RESULTS/"excel_validation.txt").write_text("\n".join(lines),encoding="utf-8")
    return result


if __name__=="__main__":
    print(json.dumps(verify_all(write=True),ensure_ascii=False,indent=2))
