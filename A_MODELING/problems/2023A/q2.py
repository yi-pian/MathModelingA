"""Question 2: parameterized common-size heliostat-field design."""

from __future__ import annotations

import json
import sys
from copy import copy
from pathlib import Path
from time import perf_counter

import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from core.export import fill_excel_template, write_excel_checked
from field import aggregate_monthly_and_annual, evaluate_representative_year, evaluate_time
from layout import constraint_report, generate_hexagonal_layout
from problem_data import OFFICIAL_DIR, RESULT_DIR
from solar import representative_times


SCREEN_MONTHS = {1, 4, 7, 10}
SCREEN_HOURS = {9.0, 12.0, 15.0}


def screen_design(design, precision="FAST"):
    summaries = []
    for month, local_time, _, _, _, sun, dni in representative_times():
        if month in SCREEN_MONTHS and local_time in SCREEN_HOURS:
            result = evaluate_time(design, sun, dni, precision=precision)
            summaries.append(result.summary(design.areas))
    frame = pd.DataFrame(summaries)
    return {key: float(frame[key].mean()) for key in frame.columns}


def candidate_grid():
    """Small transparent design grid; each candidate retains all optical terms."""
    cases = [(50.0, size, size) for size in (6.0, 6.2, 6.4, 6.5, 6.6, 6.7, 7.0, 7.4)]
    cases += [(tower_y, 8.0, 8.0) for tower_y in (-75.0, 0.0, 75.0, 150.0)]
    for tower_y, width, height in cases:
        yield {
            "tower_y_m": tower_y,
            "width_m": width,
            "height_m": height,
            "installation_height_m": max(4.0, height / 2.0 + 0.05),
            "spacing_gap_m": 0.05,
            "lattice_angle_rad": 0.0,
        }


def construct(parameters):
    return generate_hexagonal_layout(
        np.array([0.0, parameters["tower_y_m"]]),
        parameters["width_m"],
        parameters["height_m"],
        parameters["installation_height_m"],
        spacing_gap=parameters["spacing_gap_m"],
        lattice_angle_rad=parameters["lattice_angle_rad"],
    )


def write_official_result(design, path):
    values = {}
    for row, index in enumerate(range(len(design.centers)), start=2):
        values.update(
            {
                f"A{row}": float(design.tower_xy[0]),
                f"B{row}": float(design.tower_xy[1]),
                f"C{row}": index + 1,
                f"D{row}": float(design.widths[index]),
                f"E{row}": float(design.heights[index]),
                f"F{row}": float(design.centers[index, 0]),
                f"G{row}": float(design.centers[index, 1]),
                f"H{row}": float(design.centers[index, 2]),
            }
        )
    fill_excel_template(OFFICIAL_DIR / "result2.xlsx", path, "Sheet1", values)
    extend_template_row_styles(path, "Sheet1", template_row=7, final_row=len(design.centers) + 1)
    frame = pd.read_excel(path)
    if frame.shape != (len(design.centers), 8) or frame.iloc[:, 2].tolist() != list(range(1, len(design.centers) + 1)):
        raise RuntimeError("result2.xlsx row count or mirror ordering failed verification")
    numeric = frame.to_numpy(float)
    if np.any(~np.isfinite(numeric)):
        raise RuntimeError("result2.xlsx contains NaN/Inf in mirror data")
    return {"shape": frame.shape, "columns": frame.columns.tolist(), "nan_count_mirror_data": int(np.isnan(numeric).sum())}


def extend_template_row_styles(path, sheet_name, *, template_row, final_row):
    """Repeat the official example-row style without changing values or columns."""
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    for row in range(template_row + 1, final_row + 1):
        worksheet.row_dimensions[row].height = worksheet.row_dimensions[template_row].height
        for column in range(1, worksheet.max_column + 1):
            source = worksheet.cell(template_row, column)
            target = worksheet.cell(row, column)
            target._style = copy(source._style)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
    workbook.save(path)


def run():
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    started = perf_counter()
    search_path = RESULT_DIR / "q2_optimization_search.csv"
    if search_path.exists():
        search = pd.read_csv(search_path)
    else:
        search_rows = []
        for number, parameters in enumerate(candidate_grid(), start=1):
            design = construct(parameters)
            constraints = constraint_report(design)
            screened = screen_design(design)
            search_rows.append({"candidate": number, **parameters, **constraints, **{f"screen_{k}": v for k, v in screened.items()}})
            print(number, parameters, screened["power_kw"] / 1000.0, screened["power_per_area_kw_m2"])
        search = pd.DataFrame(search_rows)
    # Three FINAL calibration candidates locate the discontinuous lattice boundary; 60.35 MW
    # in this fixed 12-point screen retained a positive FINAL margin for the selected family.
    feasible = search[(search["screen_power_kw"] >= 60_350.0) & search["all_geometric_constraints_pass"]]
    if feasible.empty:
        raise RuntimeError("no screened Q2 design reaches the conservative rated-power threshold")
    best_row = feasible.sort_values("screen_power_per_area_kw_m2", ascending=False).iloc[0]
    keys = ["tower_y_m", "width_m", "height_m", "installation_height_m", "spacing_gap_m", "lattice_angle_rad"]
    parameters = {key: float(best_row[key]) for key in keys}
    height_rows = []
    for installation_height in (3.30, 3.50, 3.75, 4.00):
        candidate = dict(parameters)
        candidate["installation_height_m"] = installation_height
        screened = screen_design(construct(candidate))
        height_rows.append({"installation_height_m": installation_height, **screened})
    height_slice = pd.DataFrame(height_rows)
    best_height = height_slice.sort_values("power_per_area_kw_m2", ascending=False).iloc[0]
    parameters["installation_height_m"] = float(best_height["installation_height_m"])
    design = construct(parameters)
    cached_final = RESULT_DIR / "q2_candidate_w6p5_z3p3_final_times.csv"
    if parameters == {
        "tower_y_m": 50.0,
        "width_m": 6.5,
        "height_m": 6.5,
        "installation_height_m": 3.3,
        "spacing_gap_m": 0.05,
        "lattice_angle_rad": 0.0,
    } and cached_final.exists():
        time_frame = pd.read_csv(cached_final)
    else:
        time_frame, _, _ = evaluate_representative_year(design, precision="FINAL")
    monthly, annual = aggregate_monthly_and_annual(time_frame)
    constraints = constraint_report(design, rated_power_mw=annual["power_mw"])
    if not constraints["all_geometric_constraints_pass"] or annual["power_mw"] < 60.0:
        raise RuntimeError(f"Q2 FINAL design violates constraints: {constraints}")

    search.to_csv(RESULT_DIR / "q2_optimization_search.csv", index=False, encoding="utf-8-sig")
    height_slice.to_csv(RESULT_DIR / "q2_height_slice.csv", index=False, encoding="utf-8-sig")
    time_frame.to_csv(RESULT_DIR / "q2_time_results.csv", index=False, encoding="utf-8-sig")
    monthly.to_csv(RESULT_DIR / "q2_monthly_results.csv", index=False, encoding="utf-8-sig")
    annual_frame = pd.DataFrame([{**annual, **parameters, **constraints}])
    write_excel_checked(RESULT_DIR / "q2_tables.xlsx", {"表1_月均": monthly, "表2_年均": annual_frame, "优化搜索": search, "高度切片": height_slice}, decimals=6)
    excel_check = write_official_result(design, RESULT_DIR / "result2.xlsx")
    report = {
        "parameters": parameters,
        "annual": annual,
        "constraints": constraints,
        "screen_value": {
            "power_mw": float(best_row["screen_power_kw"] / 1000.0),
            "power_per_area_kw_m2": float(best_row["screen_power_per_area_kw_m2"]),
        },
        "final_verified_value": {
            "power_mw": annual["power_mw"],
            "power_per_area_kw_m2": annual["power_per_area_kw_m2"],
        },
        "evaluations": len(search),
        "elapsed_seconds": perf_counter() - started,
        "excel": excel_check,
        "height_slice": height_slice.to_dict(orient="records"),
        "model_note": "clipped triangular-lattice parameterization; MODEL_CONFIRMATION_REQUIRED",
    }
    (RESULT_DIR / "q2_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return design, time_frame, monthly, annual, search, report


if __name__ == "__main__":
    run()
